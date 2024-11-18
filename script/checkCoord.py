import openpyxl


def read_excel():
    wb = openpyxl.load_workbook("天宝区域1(1).xlsx")
    ws = wb.active
    dict_line = {}
    # 读取B列D列F列，B为键，D和F为值,逗号分隔
    # B为键，D和F为值,逗号分隔 可能匹配多行的值
    for row in ws.iter_rows(min_row=1, min_col=2, max_col=2):
        for cell in row:
            key = cell.value
            if key in dict_line:
                lon = ws.cell(row=cell.row, column=6).value
                lat = ws.cell(row=cell.row, column=4).value
                dict_line[key].append(str(lon) + ',' + str(lat))
            else:
                dict_line[key] = []
    print(dict_line)


if __name__ == '__main__':
    read_excel()
