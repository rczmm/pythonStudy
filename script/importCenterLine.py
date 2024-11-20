import math
import random
import requests
import openpyxl
import pyproj

# 开发者密钥
ak = "uUFjd9j2LJsDSdQxZnjYjdwWuYRQvFVz"

x_PI = 3.14159265358979324 * 3000.0 / 180.0
PI = 3.1415926535897932384626
a = 6378245.0
ee = 0.00669342162296594323


def direction_lite(origin_arg, destination_arg, waypoints_arg):
    # 倒转经纬度
    origin_arg = ','.join(reversed(origin_arg.split(',')))
    destination_arg = ','.join(reversed(destination_arg.split(',')))
    waypoints_parts = waypoints_arg.split('|')
    waypoints_arg = '|'.join([','.join(reversed(part.split(','))) for part in waypoints_parts])

    params = {
        'origin': origin_arg,
        'destination': destination_arg,
        'waypoints': waypoints_arg,
        'ak': ak
    }
    url = "https://api.map.baidu.com/directionlite/v1/driving"
    response = requests.get(url, params=params)
    response.raise_for_status()  # 抛出HTTP错误
    return response.json()


def transform_lat(lng_arg, lat_arg):
    ret = (-100.0 + 2.0 * lng_arg + 3.0 * lat_arg + 0.2 * lat_arg * lat_arg + 0.1 * lng_arg * lat_arg + 0.2 * math.sqrt(
        math.fabs(lng_arg)))
    ret += (20.0 * math.sin(6.0 * lng_arg * PI) + 20.0 * math.sin(2.0 * lng_arg * PI)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat_arg * PI) + 40.0 * math.sin(lat_arg / 3.0 * PI)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat_arg / 12.0 * PI) + 320 * math.sin(lat_arg * PI / 30.0)) * 2.0 / 3.0
    return ret


def transform_lng(lng_arg, lat_arg):
    ret = (300.0 + lng_arg + 2.0 * lat_arg + 0.1 * lng_arg * lng_arg + 0.1 * lng_arg * lat_arg + 0.1 * math.sqrt(
        math.fabs(lng_arg)))
    ret += (20.0 * math.sin(6.0 * lng_arg * PI) + 20.0 * math.sin(2.0 * lng_arg * PI)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng_arg * PI) + 40.0 * math.sin(lng_arg / 3.0 * PI)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng_arg / 12.0 * PI) + 300.0 * math.sin(lng_arg / 30.0 * PI)) * 2.0 / 3.0
    return ret


def bd09_to_gcj02(bd_lon, bd_lat):
    x = bd_lon - 0.0065
    y = bd_lat - 0.006
    z = math.sqrt(x * x + y * y) - 0.00002 * math.sin(y * x_PI)
    theta = math.atan2(y, x) - 0.000003 * math.cos(x * x_PI)
    gg_lng = z * math.cos(theta)
    gg_lat = z * math.sin(theta)
    return gg_lng, gg_lat


def gcj02_to_wgs84(lng_arg, lat_arg):
    dlat = transform_lat(lng_arg - 105.0, lat_arg - 35.0)
    dlng = transform_lng(lng_arg - 105.0, lat_arg - 35.0)
    radlat = lat_arg / 180.0 * PI
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * PI)
    dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * PI)
    mglat = lat_arg + dlat
    mglng = lng_arg + dlng
    return mglng, mglat


def wgs84_to_utm(wgs84_lon, wgs84_lat):
    utm_zone = int((wgs84_lon + 180) / 6) + 1
    wgs84 = pyproj.Proj(proj='latlong', datum='WGS84')
    utm = pyproj.Proj(proj='utm', zone=utm_zone, datum='WGS84')
    utm_east, utm_north = pyproj.transform(wgs84, utm, wgs84_lon, wgs84_lat)
    return utm_east, utm_north, utm_zone


def extract_waypoints(result_arg):
    waypoints_arg = []
    for route in result_arg['result']['routes']:
        for step in route['steps']:
            path_points = step['path'].split(';')
            for point in path_points:
                lng, lat = map(float, point.split(','))
                waypoints_arg.append({'lng': lng, 'lat': lat})
    return waypoints_arg


def write_to_excel(waypoints_arg, output_file):
    wb = openpyxl.load_workbook("中心线信息导入模板.xlsx")
    ws = wb.active
    row = 6
    i = 1
    for wp in waypoints_arg:
        ws[f'A{row}'] = f'JHC_TEST_000{i}'
        ws[f'B{row}'] = 'BAIDU'
        ws[f'C{row}'] = 'WGS84'
        gcj_lon, gcj_lat = bd09_to_gcj02(wp['lng'], wp['lat'])
        wgs_lon, wgs_lat = gcj02_to_wgs84(gcj_lon, gcj_lat)
        utm_east, utm_north, utm_zone = wgs84_to_utm(wgs_lon, wgs_lat)
        ws[f'D{row}'] = utm_east
        ws[f'E{row}'] = utm_north
        ws[f'F{row}'] = utm_zone
        ws[f'G{row}'] = wgs_lat
        ws[f'H{row}'] = wgs_lon
        ws[f'I{row}'] = random.randint(3, 100)
        ws[f'J{row}'] = random.randint(0, 200)
        row += 1
        i += 1
    wb.save(output_file)


if __name__ == '__main__':
    try:
        # 输入起点、终点和途经点
        origin = input("请输入起点坐标(经度在前，纬度在后): ")
        destination = input("请输入终点坐标(经度在前，纬度在后): ")
        waypoints = input("请输入途经点坐标(经度在前，纬度在后，坐标点用|英文竖线隔开): ")

        # 百度路径规划
        result = direction_lite(origin, destination, waypoints)

        # 提取所有的途径点
        waypoints = extract_waypoints(result)

        # 写入Excel文件
        write_to_excel(waypoints, '中心线信息导入模板.xlsx')

        print("数据已成功写入Excel文件")

        # 等待用户输入
        input("按任意键继续...")

    except requests.RequestException as e:
        print(f"请求失败: {e}")
    except KeyError as e:
        print(f"数据解析错误: {e}")
    except Exception as e:
        print(f"发生错误: {e}")
