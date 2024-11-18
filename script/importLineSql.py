import csv

import openpyxl
import requests

# 开发者密钥
ak = "uUFjd9j2LJsDSdQxZnjYjdwWuYRQvFVz"


def geo_conv(coords):
    # 接口地址
    url = "https://api.map.baidu.com/geoconv/v2/"
    # gps 转 09 model:2
    params = {
        "coords": coords,
        "model": "2",
        "ak": ak,
    }
    response = requests.get(url=url, params=params)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"请求失败，状态码: {response.status_code}")


def read_excel_columns():
    # 加载工作簿
    wb = openpyxl.load_workbook("sql(1).xlsx")

    # 获取活动工作表
    ws = wb.active

    # 初始化一个空列表来存储数据
    data = []
    coords = []

    # 遍历每一行
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=15):
        # 提取A到O列的数据
        row_data = [cell.value for cell in row]
        # print(row_data)
        row_data[0] = '''
        insert "JHC-MH-GIS"."gis_centerline"(line_id,position_x,position_y,position_b,position_l,is_deleted,tenant_id,gcj_b,gcj_l) values(
        '''
        coords.append(f'{row_data[7]},{row_data[9]}')
        data.append(row_data)

    return data, coords


def process_coords_in_batches(coords, batch_size=100):
    res = []
    # 每 200 个一组处理坐标
    for i in range(0, len(coords), batch_size):
        batch = coords[i:i + batch_size]
        coords_str = ';'.join(batch)
        json = geo_conv(coords_str)
        res.append(json)
    return res


def generate_sql_statements(data, converted_coords):
    sql_statements = []
    for i, row_data in enumerate(data):
        if i < len(converted_coords):
            coords = converted_coords[i]

            coords = coords['result']
            for coord in coords:
                sql_statement = f'''
                insert "JHC-MH-GIS"."gis_centerline"(line_id,position_x,position_y,position_b,position_l,is_deleted,tenant_id,gcj_b,gcj_l) values(
                {row_data[1]}, {row_data[3]}, {row_data[5]}, {row_data[7]}, {row_data[9]}, {row_data[11]}, '{row_data[13]}', {coord['x']}, {coord['y']}
                );
                '''
                sql_statements.append(sql_statement.strip())
    return sql_statements


def write_to_excel(sql_statements):
    # 创建新的工作簿
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # 写入数据
    for i, sql_statement in enumerate(sql_statements):
        new_ws.cell(row=i + 1, column=1, value=sql_statement)
    # 保存新的工作簿
    new_wb.save("new_sql.xlsx")


def write_sql_statements_to_csv(sql_statements, output_file="new_sql.csv"):
    try:
        with open(output_file, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            for sql_statement in sql_statements:
                writer.writerow([sql_statement])
        print(f"文件已成功保存到: {output_file}")
    except PermissionError as e:
        print(f"权限错误: {e}")
        print("请确保文件没有被其他程序占用，并且你有权限写入该文件。")
    except Exception as e:
        print(f"发生错误: {e}")


if __name__ == '__main__':
    data, coords = read_excel_columns()
    str_coord = process_coords_in_batches(coords)
    gs_84 = generate_sql_statements(data, str_coord)
    write_sql_statements_to_csv(gs_84)
